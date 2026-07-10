"""Excel 格式读写

依赖: pip install pandas openpyxl
Python 3.12。运行: python 03_excel.py

报表 / 多 Sheet / 有样式需求时用。读写速度远慢于其他格式，不适合批量程序处理。
"""
import pathlib

import numpy as np
import pandas as pd

DATA_DIR = pathlib.Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

rng = np.random.default_rng(42)
N = 100
df = pd.DataFrame({
    "date":     pd.date_range("2026-01-01", periods=N, freq="D"),
    "symbol":   rng.choice(["AAPL", "MSFT", "GOOG", "AMZN"], N),
    "close":    rng.uniform(100, 500, N).round(2),
    "volume":   rng.integers(100_000, 5_000_000, N),
    "category": rng.choice(["A", "B", "C"], N),
})


def demo01_basic():
    """① 单表读写"""
    print("① 单表读写")
    path = DATA_DIR / "data.xlsx"

    df.to_excel(path, sheet_name="行情", index=False, engine="openpyxl")
    print(f"  写出: {path.name}  {path.stat().st_size:,} bytes")

    back = pd.read_excel(path, sheet_name="行情", engine="openpyxl")
    print(f"  读回: {back.shape}  dtypes → close:{back['close'].dtype}")


def demo02_multi_sheet():
    """② 多 Sheet 读写"""
    print("\n② 多 Sheet")
    path = DATA_DIR / "multi.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df[df["symbol"] == "AAPL"].to_excel(writer, sheet_name="AAPL", index=False)
        df[df["symbol"] == "MSFT"].to_excel(writer, sheet_name="MSFT", index=False)
        df[df["symbol"] == "GOOG"].to_excel(writer, sheet_name="GOOG", index=False)
        df.describe().round(2).to_excel(writer, sheet_name="统计汇总")
    print(f"  写出: {path.name}  {path.stat().st_size:,} bytes")

    # 读所有 Sheet → {sheet_name: DataFrame}
    sheets = pd.read_excel(path, sheet_name=None, engine="openpyxl")
    for name, sdf in sheets.items():
        print(f"  Sheet [{name}]: {sdf.shape}")


def demo03_options():
    """③ 常用读取参数"""
    print("\n③ 读取参数")
    path = DATA_DIR / "data.xlsx"

    # 只读部分列
    part = pd.read_excel(path, usecols=["date", "close"])
    print(f"  usecols 2列: {part.shape}")

    # 跳过行 / 只读 N 行
    sub = pd.read_excel(path, skiprows=range(1, 11), nrows=20)
    print(f"  skiprows+nrows: {sub.shape}")

    # 指定列作为索引
    with_idx = pd.read_excel(path, index_col="date")
    print(f"  index_col=date: {with_idx.index.dtype}")


def demo04_style():
    """④ 写出带样式（标题加粗 / 列宽自适应）"""
    print("\n④ 带样式写出")
    try:
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  需要 openpyxl: pip install openpyxl")
        return

    path = DATA_DIR / "styled.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.head(20).to_excel(writer, sheet_name="样式示例", index=False)
        ws = writer.sheets["样式示例"]

        # 标题行加粗 + 背景色
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # 列宽自适应
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    print(f"  写出带样式: {path.name}  {path.stat().st_size:,} bytes")


if __name__ == "__main__":
    demo01_basic()
    demo02_multi_sheet()
    demo03_options()
    demo04_style()
